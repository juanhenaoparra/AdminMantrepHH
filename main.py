from openpyxl import load_workbook
from tinydb import TinyDB, Query
import os

""" Método que suma el contador de archivos e instancia el siguiente metodo para crear el archivo de excel """
def CrearArchivo(client_name, tipo):
	#Conexión a BD
	db = TinyDB('db/db.json')
	Cliente = Query()

	#Guarda los elementos que pertenecen al nombre pasado por parametro
	queryClient = db.search(Cliente.Client == client_name)

	if tipo == 'c':
		Count = int(queryClient[0]['CountCot'])
		newCount = Count + 1 #Le suma 1 al contador
		db.update({'CountCot':newCount}, Cliente.Client == client_name) #Actualiza el contador en la BD
		os.system('copy HojadeCotizacion.xlsx "..\\%s"\\COTIZACION' % client_name) #Copia el archivo en la carpeta respectiva del cliente
		
		#Condicional compuesto que renombra el archivo segun su código
		if newCount <= 9:
			os.rename('..//'+client_name+'//COTIZACION//HojadeCotizacion.xlsx', '..//'+client_name+'//COTIZACION//C-0000'+str(newCount)+'.xlsx')
		elif newCount >= 10 and newCount <= 99:
			os.rename('..//'+client_name+'//COTIZACION//HojadeCotizacion.xlsx', '..//'+client_name+'//COTIZACION//C-000'+str(newCount)+'.xlsx')
		elif newCount >= 100 and newCount <= 999:
			os.rename('..//'+client_name+'//COTIZACION//HojadeCotizacion.xlsx', '..//'+client_name+'//COTIZACION//C-00'+str(newCount)+'.xlsx')
		elif newCount >= 1000 and newCount <= 9999:
			os.rename('..//'+client_name+'//COTIZACION//HojadeCotizacion.xlsx', '..//'+client_name+'//COTIZACION//C-0'+str(newCount)+'.xlsx')
		elif newCount >= 10000 and newCount <= 99999:
			os.rename('..//'+client_name+'//COTIZACION//HojadeCotizacion.xlsx', '..//'+client_name+'//COTIZACION//C-'+str(newCount)+'.xlsx')
	else:
		Count = int(queryClient[0]['CountCobro'])
		newCount = Count + 1 #Le suma 1 al contador
		db.update({'CountCobro':newCount}, Cliente.Client == client_name) #Actualiza el contador en la BD
		os.system('copy HojadeCotizacion.xlsx "..\\%s"\\COBRO' % client_name) #Copia el archivo en la carpeta respectiva del cliente

		#Condicional compuesto que renombra el archivo segun su código
		if newCount <= 9:
			os.rename('..//'+client_name+'//COBRO//HojadeCotizacion.xlsx', '..//'+client_name+'//COBRO//B-0000'+str(newCount)+'.xlsx')
		elif newCount >= 10 and newCount <= 99:
			os.rename('..//'+client_name+'//COBRO//HojadeCotizacion.xlsx', '..//'+client_name+'//COBRO//B-000'+str(newCount)+'.xlsx')
		elif newCount >= 100 and newCount <= 999:
			os.rename('..//'+client_name+'//COBRO//HojadeCotizacion.xlsx', '..//'+client_name+'//COBRO//B-00'+str(newCount)+'.xlsx')
		elif newCount >= 1000 and newCount <= 9999:
			os.rename('..//'+client_name+'//COBRO//HojadeCotizacion.xlsx', '..//'+client_name+'//COBRO//B-0'+str(newCount)+'.xlsx')
		elif newCount >= 10000 and newCount <= 99999:
			os.rename('..//'+client_name+'//COBRO//HojadeCotizacion.xlsx', '..//'+client_name+'//COBRO//B-'+str(newCount)+'.xlsx')

	#Llama la funcion para manipular el archivo creado
	ManipularArchivo(client_name, tipo)

""" Método que modifica y abre la cotizacion en Excel """
def ManipularArchivo(client_name, tipo):
	#Conexión a BD
	db = TinyDB('db/db.json')
	Cliente = Query()

	#Guarda los elementos que pertenecen al nombre pasado por parametro
	queryClient = db.search(Cliente.Client == client_name)
	RepCli = queryClient[0]['Name']
	AddCli = queryClient[0]['Address']
	CityCli = queryClient[0]['City']
	PhoneCli = queryClient[0]['Phone']
	EmailCli = queryClient[0]['Email']
	nCot = '' #String vacío que almacenará el codigo del archivo
	
	if tipo == 'c':
		Count = int(queryClient[0]['CountCot'])

		#Condicional que guarda en 'nCot' el código
		if Count <= 9:
			nCot = 'C-0000'+str(Count)
		elif Count >= 10 and Count <= 99:
			nCot = 'C-000'+str(Count)
		elif Count >= 100 and Count <= 999:
			nCot = 'C-00'+str(Count)
		elif Count >= 1000 and Count <= 9999:
			nCot = 'C-0'+str(Count)
		elif Count >= 10000 and Count <= 99999:
			nCot = 'C-'+str(Count)

		#Ruta donde se ubica el archivo del cliente
		FILE_PATH = '..//'+client_name+'//COTIZACION//'+nCot+'.xlsx'
		SHEET = 'Tabla'

		#Se usa el modulo openpyxl
		workbook = load_workbook(FILE_PATH)
		sheet = workbook[SHEET]

		comandoInicio = 'start ..//"%s"//COTIZACION//%s.xlsx' % (client_name, nCot)

	else:
		Count = int(queryClient[0]['CountCobro'])

		#Condicional que guarda en 'nCot' el código
		if Count <= 9:
			nCot = 'B-0000'+str(Count)
		elif Count >= 10 and Count <= 99:
			nCot = 'B-000'+str(Count)
		elif Count >= 100 and Count <= 999:
			nCot = 'B-00'+str(Count)
		elif Count >= 1000 and Count <= 9999:
			nCot = 'B-0'+str(Count)
		elif Count >= 10000 and Count <= 99999:
			nCot = 'B-'+str(Count)

		#Ruta donde se ubica el archivo del cliente
		FILE_PATH = '..//'+client_name+'//COBRO//'+nCot+'.xlsx'
		SHEET = 'Tabla'

		#Se usa el modulo openpyxl
		workbook = load_workbook(FILE_PATH)
		sheet = workbook[SHEET]

		sheet['D1'] = 'CUENTA DE COBRO'
		comandoInicio = 'start ..//"%s"//COBRO//%s.xlsx' % (client_name, nCot)
	
	#Se modifican las celdas con los datos del cliente en la BD
	sheet['B4'] = client_name
	sheet['B6'] = nCot
	sheet['B10'] = AddCli
	sheet['B12'] = CityCli
	sheet['B14'] = PhoneCli
	sheet['B16'] = EmailCli
	sheet['B18'] = RepCli
	
	#Se guardan los cambios hechos
	workbook.save(FILE_PATH)

	#Se abre el archivo
	os.system(comandoInicio)
